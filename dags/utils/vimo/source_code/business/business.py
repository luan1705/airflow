import pandas as pd
import re
import os
import glob
from datetime import date
from sqlalchemy import create_engine, text
import openpyxl

DB_URL = "postgresql+psycopg2://root:Dnl_123456@tanhungsoft.com:5432/dnl"
engine = create_engine(DB_URL)

SCHEMA = "macro"
TABLE  = "business"

COLS = [
    "newBusiness",
    "returnBusiness",
    "exitBusiness",
    "newBusiness_yoy",   # script cal tính sau
    "exitBusiness_yoy",  # script cal tính sau
]

ROWS = {
    "Doanh nghiệp đăng ký thành lập mới":     "newBusiness",
    "Doanh nghiệp quay trở lại hoạt động":    "returnBusiness",
    "Doanh nghiệp tạm ngừng kinh doanh":      "exit_1",
    "Doanh nghiệp tạm ngừng hoạt động":       "exit_2",
    "Doanh nghiệp hoàn tất thủ tục giải thể": "exit_3",
}


def parse_time_from_filename(file_path: str) -> date:
    name = file_path.split("/")[-1].split("\\")[-1]
    match = re.search(r'(\d{4})_(\d{2})', name)
    if not match:
        raise ValueError(f"Không parse được tháng/năm từ tên file: {name}")
    return date(int(match.group(1)), int(match.group(2)), 1)


def get_sheet_name(file_path: str) -> str:
    xl = pd.ExcelFile(file_path)
    for sheet in xl.sheet_names:
        s = sheet.lower().replace(' ', '').replace('.', '')
        if ('chitieu' in s or 'chỉtieudn' in s or 'chỉtiêudn' in s) and 'dn' in s:
            return sheet
    for sheet in xl.sheet_names:
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        ws = wb[sheet]
        content = ' '.join([str(v) for row in ws.iter_rows(max_row=5, values_only=True) for v in row if v])
        wb.close()
        if 'chỉ tiêu' in content.lower() and 'doanh nghiệp' in content.lower():
            return sheet
    raise ValueError(f"Không tìm thấy sheet Chi tieu DN trong {file_path}")


def to_float(val):
    try:
        return float(val)
    except (TypeError, ValueError):
        return None


def get_value_col(ws, year: int):
    col_texts = {}
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if any(v is not None for v in row):
            for ci, val in enumerate(row):
                if val is not None:
                    col_texts[ci] = col_texts.get(ci, '') + ' ' + str(val).strip().lower()
        if i > 10:
            break

    month_cols = []
    for ci, text in sorted(col_texts.items()):
        if ci <= 0:
            continue
        if str(year) in text and 'tháng' in text and 'so với' not in text and '%' not in text:
            if not re.search(r'[0-9]+ tháng', text):
                month_cols.append(ci)

    return month_cols[-1] if month_cols else None


def parse_business(file_path: str) -> pd.DataFrame:
    time  = parse_time_from_filename(file_path)
    year  = time.year
    sheet = get_sheet_name(file_path)

    wb      = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    ws      = wb[sheet]
    val_col = get_value_col(ws, year)

    raw = {}
    for row in ws.iter_rows(values_only=True):
        if not row[0]:
            continue
        name      = str(row[0]).strip().replace('\n', ' ')
        name_norm = ' '.join(name.split())
        for key in ROWS:
            if name_norm.startswith(key):
                col_key = ROWS[key]
                if col_key not in raw:
                    raw[col_key] = to_float(row[val_col] if val_col and len(row) > val_col else None)
                break

    wb.close()

    exit_vals  = [raw.get(f'exit_{i}') for i in [1, 2, 3]]
    exit_total = sum(v for v in exit_vals if v is not None) if any(v is not None for v in exit_vals) else None

    result = {
        "time":              time,
        "newBusiness":       raw.get("newBusiness"),
        "returnBusiness":    raw.get("returnBusiness"),
        "exitBusiness":      exit_total,
        "newBusiness_yoy":   None,
        "exitBusiness_yoy":  None,
    }

    return pd.DataFrame([result])


def upsert_business(df: pd.DataFrame):
    # Chỉ upsert các cột không phải yoy — script cal tính sau
    upsert_cols = [c for c in COLS if '_yoy' not in c]

    with engine.begin() as conn:
        conn.execute(text(f'CREATE SCHEMA IF NOT EXISTS "{SCHEMA}"'))
        col_defs  = "\n".join([f'    "{c}" DOUBLE PRECISION,' for c in COLS[:-1]])
        col_defs += f'\n    "{COLS[-1]}" DOUBLE PRECISION'
        conn.execute(text(f"""
            CREATE TABLE IF NOT EXISTS {SCHEMA}.{TABLE} (
                time DATE PRIMARY KEY,
{col_defs}
            )
        """))
        for col in COLS:
            conn.execute(text(f"""
                ALTER TABLE {SCHEMA}.{TABLE}
                ADD COLUMN IF NOT EXISTS "{col}" DOUBLE PRECISION
            """))

        set_clause  = ",\n".join([f'    "{c}" = EXCLUDED."{c}"' for c in upsert_cols])
        insert_cols = ", ".join([f'"{c}"' for c in upsert_cols])
        insert_vals = ", ".join([f":{c}" for c in upsert_cols])
        records = df[['time'] + upsert_cols].replace({float('nan'): None}).to_dict(orient="records")
        conn.execute(text(f"""
            INSERT INTO {SCHEMA}.{TABLE} (time, {insert_cols})
            VALUES (:time, {insert_vals})
            ON CONFLICT (time) DO UPDATE SET
{set_clause}
        """), records)

    print(f"✅ Upsert {len(df)} rows vào {SCHEMA}.{TABLE}")


def save_business(file_path: str):
    df = parse_business(file_path)
    print(df.T.to_string())
    upsert_business(df)


def _sort_key(f):
    match = re.search(r'(\d{4})_(\d{2})', os.path.basename(f))
    return (int(match.group(1)), int(match.group(2))) if match else (0, 0)


def get_latest_file(data_dir: str) -> str:
    files = glob.glob(os.path.join(data_dir, "*.xlsx"))
    if not files:
        raise FileNotFoundError(f"Không tìm thấy file xlsx trong {data_dir}")
    return sorted(files, key=_sort_key)[-1]


#=======================Chạy file chỉ định trực tiếp trong terminal=====================
# def business(**context):
#     save_business("../../data/2026_01.xlsx")

#=======================Chạy file chỉ định airflow=====================
# def business(**context):
#     save_business("/opt/airflow/dags/utils/vimo/data/2026_01.xlsx")

# =====================Chạy file mới nhất=====================
# def business(**context):
#     data_dir  = os.path.join(os.path.dirname(__file__), "../../data")
#     file_path = get_latest_file(data_dir)
#     print(f"📂 File mới nhất: {file_path}")
#     save_business(file_path)

# =====================Chạy tất cả file=====================
def business(**context):
    data_dir = os.path.join(os.path.dirname(__file__), "../../data")
    files    = glob.glob(os.path.join(data_dir, "*.xlsx"))
    if not files:
        raise FileNotFoundError(f"Không tìm thấy file xlsx trong {data_dir}")
    for file_path in sorted(files, key=_sort_key):
        print(f"📂 Đang chạy: {file_path}")
        save_business(file_path)

#===============================================================

if __name__ == "__main__":
    business()