import pandas as pd
import re
import os
import glob
from datetime import date
from sqlalchemy import create_engine, text
import openpyxl

DB_URL = "postgresql+psycopg2://vnsfintech:Vns_123456@tanhungsoft.com:5433/vnsfintech"
engine = create_engine(DB_URL)

SCHEMA = "macro"
TABLE  = "transport_tourism"


def parse_time_from_filename(file_path: str) -> date:
    name = file_path.split("/")[-1].split("\\")[-1]
    match = re.search(r'(\d{4})_(\d{2})', name)
    if not match:
        raise ValueError(f"Không parse được tháng/năm từ tên file: {name}")
    return date(int(match.group(1)), int(match.group(2)), 1)


def get_hk_sheet(file_path: str) -> str:
    xl = pd.ExcelFile(file_path)
    for sheet in xl.sheet_names:
        s = sheet.lower().replace(' ', '').replace('.', '')
        if ('vantaihk' in s or 'vthk' in s) and 'quy' not in s:
            return sheet
    for sheet in xl.sheet_names:
        if re.match(r'^\s*17[\.\s]', sheet):
            return sheet
    raise ValueError(f"Không tìm thấy sheet VT HK trong {file_path}")


def get_kqt_sheet(file_path: str) -> str:
    xl = pd.ExcelFile(file_path)
    for sheet in xl.sheet_names:
        s = sheet.lower().replace(' ', '').replace('.', '')
        if 'kqt' in s or 'dulich' in s or 'khachqt' in s or 'khachquoc' in s:
            return sheet
    for sheet in xl.sheet_names:
        if re.match(r'^\s*19[\.\s]', sheet):
            return sheet
    raise ValueError(f"Không tìm thấy sheet KQT trong {file_path}")


def to_float(val):
    try:
        return float(val)
    except (TypeError, ValueError):
        return None


def detect_kqt_cols(ws, year: int):
    """Detect cột value, ytd, yoy cho sheet KQT."""
    col_texts = {}
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if any(v is not None for v in row):
            for ci, val in enumerate(row):
                if val is not None:
                    col_texts[ci] = col_texts.get(ci, '') + ' ' + str(val).strip().lower()
        if i > 10:
            break

    val_col = ytd_col = yoy_col = None

    # Value: cột "ước tính" + năm hiện tại + không phải YTD
    uoc_tinh_cols = []
    for ci, text in sorted(col_texts.items()):
        if ci <= 0:
            continue
        if 'ước tính' in text and str(year) in text and 'so với' not in text:
            if not re.search(r'[0-9]+ tháng', text) and 'quý' not in text and 'năm ' + str(year) + ' so' not in text:
                uoc_tinh_cols.append(ci)

    if uoc_tinh_cols:
        # Lấy cột ước tính tháng hiện tại (cột nhỏ nhất trong uoc_tinh_cols - tháng trước cột YTD)
        # Thực ra cần cột có "tháng X" không phải "N tháng"
        for ci in uoc_tinh_cols:
            text = col_texts.get(ci, '')
            if 'tháng' in text and not re.search(r'[0-9]+ tháng', text) and 'quý' not in text:
                val_col = ci
                break
        if val_col is None:
            val_col = uoc_tinh_cols[0]

    # Fallback: cột có "tháng X năm Y" không phải thực hiện, không so sánh
    if val_col is None:
        month_cols = []
        for ci, text in sorted(col_texts.items()):
            if ci <= 0:
                continue
            if str(year) in text and 'so với' not in text and '%' not in text and 'tháng' in text:
                if not re.search(r'[0-9]+ tháng', text) and 'quý' not in text:
                    month_cols.append(ci)
        if month_cols:
            val_col = month_cols[-1]  # Cột cuối = tháng hiện tại

    # YTD: "N tháng" hoặc "quý I/II/III" hoặc "năm YYYY" không so sánh
    for ci, text in sorted(col_texts.items()):
        if ci <= 0:
            continue
        if str(year) in text and 'so với' not in text and '%' not in text:
            if re.search(r'[0-9]+ tháng', text) or 'quý' in text or (f'năm {year}' in text and 'tháng' not in text):
                if ytd_col is None:
                    ytd_col = ci

    # YoY: "cùng kỳ" hoặc "năm trước" + so với + không phải YTD YoY (ưu tiên cột không có "N tháng/quý")
    for ci, text in sorted(col_texts.items()):
        if ci <= 0:
            continue
        if 'so với' in text and ('cùng kỳ' in text or 'năm trước' in text):
            if not re.search(r'[0-9]+ tháng', text) and 'quý' not in text:
                if yoy_col is None:
                    yoy_col = ci

    return val_col, ytd_col, yoy_col


def detect_hk_cols(ws, year: int):
    """Detect cột value và ytd cho sheet VT HK."""
    col_texts = {}
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if any(v is not None for v in row):
            for ci, val in enumerate(row):
                if val is not None:
                    col_texts[ci] = col_texts.get(ci, '') + ' ' + str(val).strip().lower()
        if i > 10:
            break

    val_col = ytd_col = None

    for ci, text in sorted(col_texts.items()):
        if ci <= 0:
            continue
        if str(year) in text and 'so với' not in text and '%' not in text:
            if re.search(r'[0-9]+ tháng', text) or (f'năm {year}' in text and 'tháng' not in text):
                if ytd_col is None:
                    ytd_col = ci
            elif 'tháng' in text and val_col is None:
                val_col = ci

    # Fallback: lấy cột cuối có năm hiện tại trước cột YTD
    if val_col is None:
        candidates = [ci for ci, text in col_texts.items() if str(year) in text and 'so với' not in text and '%' not in text]
        if ytd_col:
            candidates = [ci for ci in candidates if ci < ytd_col]
        if candidates:
            val_col = max(candidates)

    return val_col, ytd_col


def parse_transport_tourism(file_path: str) -> pd.DataFrame:
    time = parse_time_from_filename(file_path)
    year = time.year
    result = {"time": time}

    # ==================== VT HK ====================
    sheet_hk = get_hk_sheet(file_path)
    wb = openpyxl.load_workbook(file_path, read_only=True)
    ws = wb[sheet_hk]
    val_col, ytd_col = detect_hk_cols(ws, year)

    domestic_val = domestic_ytd = None
    foreign_val  = foreign_ytd  = None

    for row in ws.iter_rows(values_only=True):
        name = None
        for ci in [0, 1]:
            if len(row) > ci and row[ci] and str(row[ci]).strip() in ['Trong nước', 'Ngoài nước']:
                name = str(row[ci]).strip()
                break
        if name == 'Trong nước' and domestic_val is None:
            domestic_val = to_float(row[val_col] if val_col is not None and len(row) > val_col else None)
            domestic_ytd = to_float(row[ytd_col] if ytd_col is not None and len(row) > ytd_col else None)
        elif name == 'Ngoài nước' and foreign_val is None:
            foreign_val = to_float(row[val_col] if val_col is not None and len(row) > val_col else None)
            foreign_ytd = to_float(row[ytd_col] if ytd_col is not None and len(row) > ytd_col else None)

    wb.close()

    # T1: ytd = value
    if domestic_ytd is None:
        domestic_ytd = domestic_val
    if foreign_ytd is None:
        foreign_ytd = foreign_val

    result['transport_domestic_value'] = domestic_val
    result['transport_domestic_ytd']   = domestic_ytd
    result['transport_foreign_value']  = foreign_val
    result['transport_foreign_ytd']    = foreign_ytd
    result['transport_total_value']    = (domestic_val or 0) + (foreign_val or 0) if domestic_val or foreign_val else None
    result['transport_total_ytd']      = (domestic_ytd or 0) + (foreign_ytd or 0) if domestic_ytd or foreign_ytd else None

    # ==================== KQT ====================
    try:
        sheet_kqt = get_kqt_sheet(file_path)
    except ValueError:
        sheet_kqt = None

    if sheet_kqt:
        wb = openpyxl.load_workbook(file_path, read_only=True)
        ws = wb[sheet_kqt]
        val_col, ytd_col, yoy_col = detect_kqt_cols(ws, year)

        for row in ws.iter_rows(values_only=True):
            if row[0] and 'TỔNG SỐ' in str(row[0]):
                result['tourism_value'] = to_float(row[val_col] if val_col is not None and len(row) > val_col else None)
                result['tourism_ytd']   = to_float(row[ytd_col] if ytd_col is not None and len(row) > ytd_col else None)
                v_yoy = to_float(row[yoy_col] if yoy_col is not None and len(row) > yoy_col else None)
                result['tourism_yoy']   = round((v_yoy - 100) / 100, 4) if v_yoy is not None else None
                break
        wb.close()
    else:
        result['tourism_value'] = result['tourism_ytd'] = result['tourism_yoy'] = None

    if result.get('tourism_ytd') is None:
        result['tourism_ytd'] = result.get('tourism_value')

    return pd.DataFrame([result])


def upsert_transport_tourism(df: pd.DataFrame):
    cols = [c for c in df.columns if c != 'time']

    with engine.begin() as conn:
        conn.execute(text(f'CREATE SCHEMA IF NOT EXISTS "{SCHEMA}"'))
        col_defs  = "\n".join([f"    {c} DOUBLE PRECISION," for c in cols[:-1]])
        col_defs += f"\n    {cols[-1]} DOUBLE PRECISION"
        conn.execute(text(f"""
            CREATE TABLE IF NOT EXISTS {SCHEMA}.{TABLE} (
                time DATE PRIMARY KEY,
{col_defs}
            )
        """))
        for col in cols:
            conn.execute(text(f"""
                ALTER TABLE {SCHEMA}.{TABLE}
                ADD COLUMN IF NOT EXISTS {col} DOUBLE PRECISION
            """))

        set_clause  = ",\n".join([f"    {c} = EXCLUDED.{c}" for c in cols])
        insert_cols = ", ".join(cols)
        insert_vals = ", ".join([f":{c}" for c in cols])
        conn.execute(text(f"""
            INSERT INTO {SCHEMA}.{TABLE} (time, {insert_cols})
            VALUES (:time, {insert_vals})
            ON CONFLICT (time) DO UPDATE SET
{set_clause}
        """), df.replace({float('nan'): None}).to_dict(orient="records"))

    print(f"✅ Upsert {len(df)} rows vào {SCHEMA}.{TABLE}")


def save_transport_tourism(file_path: str):
    df = parse_transport_tourism(file_path)
    print(df.T.to_string())
    upsert_transport_tourism(df)


def _sort_key(f):
    match = re.search(r'(\d{4})_(\d{2})', os.path.basename(f))
    return (int(match.group(1)), int(match.group(2))) if match else (0, 0)


def get_latest_file(data_dir: str) -> str:
    files = glob.glob(os.path.join(data_dir, "*.xlsx"))
    if not files:
        raise FileNotFoundError(f"Không tìm thấy file xlsx trong {data_dir}")
    return sorted(files, key=_sort_key)[-1]


#=======================Chạy file chỉ định trực tiếp trong terminal=====================
# def transport_tourism(**context):
#     save_transport_tourism("../../data/2023_01.xlsx")

#=======================Chạy file chỉ định airflow=====================
# def transport_tourism(**context):
#     save_transport_tourism("/opt/airflow/dags/utils/vimo/data/2026_01.xlsx")

# =====================Chạy file mới nhất=====================
# def transport_tourism(**context):
#     data_dir  = os.path.join(os.path.dirname(__file__), "../../data")
#     file_path = get_latest_file(data_dir)
#     print(f"📂 File mới nhất: {file_path}")
#     save_transport_tourism(file_path)

# =====================Chạy tất cả file=====================
def transport_tourism(**context):
    data_dir = os.path.join(os.path.dirname(__file__), "../../data")
    files    = glob.glob(os.path.join(data_dir, "*.xlsx"))
    if not files:
        raise FileNotFoundError(f"Không tìm thấy file xlsx trong {data_dir}")
    for file_path in sorted(files, key=_sort_key):
        print(f"📂 Đang chạy: {file_path}")
        save_transport_tourism(file_path)

#===============================================================

if __name__ == "__main__":
    transport_tourism()