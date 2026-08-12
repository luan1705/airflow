import pandas as pd
import re
import os
import glob
from datetime import date
from sqlalchemy import create_engine, text
import openpyxl

DB_URL = "postgresql+psycopg2://root:Dnl_123456@tanhungsoft.com:5432/dnl"
engine = create_engine(DB_URL)

SCHEMA = "macro"
TABLE  = "cpi_mom"

COMPONENTS = {
    "CHỈ SỐ GIÁ TIÊU DÙNG":                    "cpi",
    "Hàng ăn và dịch vụ ăn uống":               "food_beverage",
    "Lương thực":                                "food_staple",
    "Thực phẩm":                                 "food",
    "Ăn uống ngoài gia đình":                    "eating_out",
    "Đồ uống và thuốc lá":                       "drink_tobacco",
    "May mặc, giày dép và mũ nón":              "clothing",
    "May mặc, mũ nón và giày dép":              "clothing",
    "May mặc, mũ nón và giày dép ":             "clothing",
    "Nhà ở, điện nước, chất đốt và VLXD(*)":   "housing",
    "Nhà ở, điện nước, chất đốt và VLXD (*)":  "housing",
    "Nhà ở, điện, nước, chất đốt và vật liệu xây dựng": "housing",
    "Nhà ở và vật liệu xây dựng":               "housing",
    "Nhà ở và vật liệu xây dựng(*)":            "housing",
    "Thiết bị và đồ dùng gia đình":             "household",
    "Thuốc và dịch vụ y tế":                    "healthcare",
    "Dịch vụ y tế":                              "medical_service",
    "Giao thông":                                "transport",
    "Bưu chính viễn thông":                      "telecom",
    "Thông tin và truyền thông":                 "telecom",
    "Giáo dục":                                  "education",
    "Dịch vụ giáo dục":                          "education_service",
    "Văn hoá, giải trí và du lịch":             "culture_entertainment",
    "Đồ dùng và dịch vụ khác":                  "other",
    "Hàng hóa và dịch vụ khác":                 "other",
    "CHỈ SỐ GIÁ VÀNG":                          "gold",
    "CHỈ SỐ GIÁ ĐÔ LA MỸ":                     "usd",
    "LẠM PHÁT CƠ BẢN":                          "core_inflation",
}

NO_INDEX = {"core_inflation"}


def parse_time_from_filename(file_path: str) -> date:
    name = file_path.split("/")[-1].split("\\")[-1]
    match = re.search(r'(\d{4})_(\d{2})', name)
    if not match:
        raise ValueError(f"Không parse được tháng/năm từ tên file: {name}")
    return date(int(match.group(1)), int(match.group(2)), 1)


def get_sheet_name(file_path: str) -> str:
    xl = pd.ExcelFile(file_path)
    for sheet in xl.sheet_names:
        if 'cpi' in sheet.lower().replace(' ', '').replace('.', ''):
            return sheet
    raise ValueError(f"Không tìm thấy sheet CPI trong {file_path}")


def to_float(val):
    try:
        return float(val)
    except (TypeError, ValueError):
        return None


def get_mom_col(ws, report_time: date):
    """
    Tìm cột so với tháng trước.

    Ví dụ:
        file 2026_07 -> tìm "Tháng 6", năm 2026
        file 2025_05 -> tìm "Tháng 4", năm 2025
        file 2026_01 -> tìm "Tháng 12", năm 2025
    """
    if report_time.month == 1:
        previous_month = 12
        previous_year = report_time.year - 1
    else:
        previous_month = report_time.month - 1
        previous_year = report_time.year

    headers = {}

    for row_idx, row in enumerate(
        ws.iter_rows(values_only=True),
        start=1,
    ):
        if row_idx > 12:
            break

        for col_idx, value in enumerate(row):
            if value is None:
                continue

            headers[col_idx] = (
                headers.get(col_idx, "")
                + " "
                + str(value).strip().lower()
            )

    month_pattern = re.compile(
        rf"\btháng\s*0?{previous_month}\b",
        flags=re.IGNORECASE,
    )

    candidates = []

    for col_idx, header in headers.items():
        normalized = re.sub(r"\s+", " ", header)

        if not month_pattern.search(normalized):
            continue

        if str(previous_year) not in normalized:
            continue

        if "bình quân" in normalized:
            continue

        if "cùng kỳ" in normalized:
            continue

        if "kỳ gốc" in normalized:
            continue

        candidates.append(col_idx)

    if not candidates:
        raise ValueError(
            f"Không tìm thấy cột MoM tương ứng "
            f"tháng {previous_month}/{previous_year}"
        )

    mom_col = candidates[-1]

    print(
        f"✅ Cột MoM: index={mom_col}, "
        f"Excel column={openpyxl.utils.get_column_letter(mom_col + 1)}, "
        f"header={headers[mom_col]}"
    )

    return mom_col


def parse_cpi_mom(file_path: str) -> pd.DataFrame:
    time  = parse_time_from_filename(file_path)
    sheet = get_sheet_name(file_path)

    wb      = openpyxl.load_workbook(file_path, read_only=True)
    ws      = wb[sheet]
    mom_col = get_mom_col(ws,time)

    unique_cols = list(dict.fromkeys(COMPONENTS.values()))
    result = {"time": time}
    for col in unique_cols:
        result[col] = None

    if mom_col is None:
        wb.close()
        return pd.DataFrame([result])

    for row in ws.iter_rows(values_only=True):
        name = None
        for ci in [2, 1, 0]:
            if len(row) > ci and row[ci] and str(row[ci]).strip():
                candidate = str(row[ci]).strip()
                if "Trong đó" not in candidate and candidate in COMPONENTS:
                    name = candidate
                    break

        if name not in COMPONENTS:
            continue

        col = COMPONENTS[name]
        if result[col] is not None:
            continue

        val = to_float(row[mom_col] if len(row) > mom_col else None)

        if col in NO_INDEX:
            result[col] = round(val/100, 6) if val is not None else None
        else:
            result[col] = round((val - 100) / 100, 6) if val is not None else None

    wb.close()
    return pd.DataFrame([result])


def upsert_cpi_mom(df: pd.DataFrame):
    cols = list(dict.fromkeys(COMPONENTS.values()))

    with engine.begin() as conn:
        conn.execute(text(f'CREATE SCHEMA IF NOT EXISTS "{SCHEMA}"'))
        col_defs  = "\n".join([f"    {c} DOUBLE PRECISION," for c in cols[:-1]])
        col_defs += f"\n    {cols[-1]} DOUBLE PRECISION"
        conn.execute(text(f"""
            CREATE TABLE IF NOT EXISTS {SCHEMA}.{TABLE} (
                time DATE PRIMARY KEY,
{col_defs}
            )
        """))
        for col in cols:
            conn.execute(text(f"""
                ALTER TABLE {SCHEMA}.{TABLE}
                ADD COLUMN IF NOT EXISTS {col} DOUBLE PRECISION
            """))

        set_clause  = ",\n".join([f"    {c} = EXCLUDED.{c}" for c in cols])
        insert_cols = ", ".join(cols)
        insert_vals = ", ".join([f":{c}" for c in cols])
        conn.execute(text(f"""
            INSERT INTO {SCHEMA}.{TABLE} (time, {insert_cols})
            VALUES (:time, {insert_vals})
            ON CONFLICT (time) DO UPDATE SET
{set_clause}
        """), df.replace({float('nan'): None}).to_dict(orient="records"))

    print(f"✅ Upsert {len(df)} rows vào {SCHEMA}.{TABLE}")


def save_cpi_mom(file_path: str):
    df = parse_cpi_mom(file_path)
    print(df.T.to_string())
    upsert_cpi_mom(df)


def _sort_key(f):
    match = re.search(r'(\d{4})_(\d{2})', os.path.basename(f))
    return (int(match.group(1)), int(match.group(2))) if match else (0, 0)


def get_latest_file(data_dir: str) -> str:
    files = glob.glob(os.path.join(data_dir, "*.xlsx"))
    if not files:
        raise FileNotFoundError(f"Không tìm thấy file xlsx trong {data_dir}")
    return sorted(files, key=_sort_key)[-1]


#=======================Chạy file chỉ định trực tiếp trong terminal=====================
# def cpi_mom(**context):
#     save_cpi_mom("../../data/excel/2023_01.xlsx")

#=======================Chạy file chỉ định airflow=====================
# def cpi_mom(**context):
#     save_cpi_mom("/opt/airflow/dags/utils/vimo/data/excel/2026_01.xlsx")

#=====================Chạy file mới nhất=====================
def cpi_mom(**context):
    data_dir  = os.path.join(os.path.dirname(__file__), "../../data/excel")
    file_path = get_latest_file(data_dir)
    print(f"📂 File mới nhất: {file_path}")
    save_cpi_mom(file_path)

# # =====================Chạy tất cả file=====================
# def cpi_mom(**context):
#     data_dir = os.path.join(os.path.dirname(__file__), "../../data/excel")
#     files    = glob.glob(os.path.join(data_dir, "*.xlsx"))
#     if not files:
#         raise FileNotFoundError(f"Không tìm thấy file xlsx trong {data_dir}")
#     for file_path in sorted(files, key=_sort_key):
#         print(f"📂 Đang chạy: {file_path}")
#         try:
#             save_cpi_mom(file_path)
#         except Exception as e:
#             print(f"⚠️ Lỗi {file_path}: {e} — upsert null")
#             try:
#                 time = parse_time_from_filename(file_path)
#                 df = pd.DataFrame([{"time": time}])
#                 upsert_cpi_mom(df)
#             except Exception as e2:
#                 print(f"⚠️ Bỏ qua {file_path}: {e2}")

#===============================================================

if __name__ == "__main__":
    cpi_mom()