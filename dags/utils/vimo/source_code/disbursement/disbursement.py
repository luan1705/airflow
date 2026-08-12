import pandas as pd
import re
import os
import glob
from datetime import date
from sqlalchemy import create_engine, text
import openpyxl

DB_URL = "postgresql+psycopg2://root:Dnl_123456@tanhungsoft.com:5432/dnl"
engine = create_engine(DB_URL)

SCHEMA = "macro"
TABLE  = "disbursement"

COLS = [
    "disbursementValue",
    "disbursementYtd",
    "disbursementPlan",
    "disbursementPlanRatio",  # disbursementCal.py tính sau
]


def parse_time_from_filename(file_path: str) -> date:
    name = file_path.split("/")[-1].split("\\")[-1]
    match = re.search(r'(\d{4})_(\d{2})', name)
    if not match:
        raise ValueError(f"Không parse được tháng/năm từ tên file: {name}")
    return date(int(match.group(1)), int(match.group(2)), 1)


def get_sheet_name(file_path: str, month: int) -> str:
    xl = pd.ExcelFile(file_path)
    for sheet in xl.sheet_names:
        if 'quy' in sheet.lower() or 'quý' in sheet.lower():
            continue
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        ws = wb[sheet]
        content = ' '.join([str(v) for row in ws.iter_rows(max_row=1, values_only=True) for v in row if v]).lower()
        wb.close()
        if 'ngân sách nhà nước' in content and 'vốn đầu tư' in content and 'toàn xã hội' not in content:
            return sheet
    raise ValueError(f"Không tìm thấy sheet VĐT NSNN trong {file_path}")


def to_float(val):
    try:
        return float(val)
    except (TypeError, ValueError):
        return None


def detect_cols(ws, year: int):
    col_texts = {}
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if any(v is not None for v in row):
            for ci, val in enumerate(row):
                if val is not None:
                    col_texts[ci] = col_texts.get(ci, '') + ' ' + str(val).strip().lower()
        if i > 8:
            break

    plan_col = value_col = ytd_col = ratio_col = None

    for ci, text in sorted(col_texts.items()):
        if ci <= 0:
            continue
        if 'kế hoạch' in text and 'so với' not in text and '%' not in text and plan_col is None:
            plan_col = ci
        if ('cộng dồn' in text or ('ước tính' in text and ('quý' in text or 'quy' in text or 'qu' in text)) or ('năm' in text and 'so với' not in text and 'kế hoạch' not in text and 'ước tính' not in text and 'thực hiện' not in text)) and str(year) in text and 'so với' not in text and '%' not in text and ytd_col is None:
            ytd_col = ci
        if 'ước tính' in text and 'tháng' in text and str(year) in text and 'qu' not in text and value_col is None:
            value_col = ci
        if 'kế hoạch' in text and ('so với' in text or '%' in text) and ratio_col is None:
            ratio_col = ci

    return plan_col, value_col, ytd_col, ratio_col


def parse_disbursement(file_path: str) -> pd.DataFrame:
    time  = parse_time_from_filename(file_path)
    year  = time.year
    month = time.month
    sheet = get_sheet_name(file_path, month)

    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    ws = wb[sheet]

    plan_col, value_col, ytd_col, ratio_col = detect_cols(ws, year)

    result = {"time": time}
    for col in COLS:
        result[col] = None

    for row in ws.iter_rows(values_only=True):
        if row[0] and 'TỔNG SỐ' in str(row[0]):
            result['disbursementPlan']      = to_float(row[plan_col]  if plan_col  and len(row) > plan_col  else None)
            result['disbursementValue']     = to_float(row[value_col] if value_col and len(row) > value_col else None)
            ytd = to_float(row[ytd_col] if ytd_col and len(row) > ytd_col else None)
            result['disbursementYtd']       = ytd if ytd is not None else result['disbursementValue']
            result['disbursementPlanRatio'] = None  # disbursementCal.py tính sau

            # Đưa về đơn vị đồng (file GSO đơn vị tỷ đồng)
            for col in ['disbursementValue', 'disbursementYtd', 'disbursementPlan']:
                if result[col] is not None:
                    result[col] = round(result[col] * 1_000_000_000)
            break

    wb.close()
    return pd.DataFrame([result])


def upsert_disbursement(df: pd.DataFrame):
    with engine.begin() as conn:
        conn.execute(text(f'CREATE SCHEMA IF NOT EXISTS "{SCHEMA}"'))
        col_defs  = "\n".join([f'    "{c}" DOUBLE PRECISION,' for c in COLS[:-1]])
        col_defs += f'\n    "{COLS[-1]}" DOUBLE PRECISION'
        conn.execute(text(f"""
            CREATE TABLE IF NOT EXISTS {SCHEMA}.{TABLE} (
                time DATE PRIMARY KEY,
{col_defs}
            )
        """))
        for col in COLS:
            conn.execute(text(f"""
                ALTER TABLE {SCHEMA}.{TABLE}
                ADD COLUMN IF NOT EXISTS "{col}" DOUBLE PRECISION
            """))

        upsert_cols = [c for c in COLS if c != 'disbursementPlanRatio']
        set_clause  = ",\n".join([f'    "{c}" = EXCLUDED."{c}"' for c in upsert_cols])
        insert_cols = ", ".join([f'"{c}"' for c in upsert_cols])
        insert_vals = ", ".join([f":{c}" for c in upsert_cols])
        records = df[['time'] + upsert_cols].replace({float('nan'): None}).to_dict(orient="records")
        conn.execute(text(f"""
            INSERT INTO {SCHEMA}.{TABLE} (time, {insert_cols})
            VALUES (:time, {insert_vals})
            ON CONFLICT (time) DO UPDATE SET
{set_clause}
        """), records)

    print(f"✅ Upsert {len(df)} rows vào {SCHEMA}.{TABLE}")


def save_disbursement(file_path: str):
    df = parse_disbursement(file_path)
    print(df.T.to_string())
    upsert_disbursement(df)


def _sort_key(f):
    match = re.search(r'(\d{4})_(\d{2})', os.path.basename(f))
    return (int(match.group(1)), int(match.group(2))) if match else (0, 0)


def get_latest_file(data_dir: str) -> str:
    files = glob.glob(os.path.join(data_dir, "*.xlsx"))
    if not files:
        raise FileNotFoundError(f"Không tìm thấy file xlsx trong {data_dir}")
    return sorted(files, key=_sort_key)[-1]


##=======================Chạy file chỉ định trực tiếp trong terminal=====================
# def disbursement(**context):
#     save_disbursement("../../data/excel/2023_01.xlsx")

# ##=======================Chạy file chỉ định airflow=====================
# def disbursement(**context):
#     save_disbursement("/opt/airflow/dags/utils/vimo/data/excel/2026_06.xlsx")

#=====================Chạy file mới nhất=====================
def disbursement(**context):
    data_dir  = os.path.join(os.path.dirname(__file__), "../../data/excel")
    file_path = get_latest_file(data_dir)
    print(f"📂 File mới nhất: {file_path}")
    save_disbursement(file_path)

# #=====================Chạy tất cả file=====================
# def disbursement(**context):
#     data_dir = os.path.join(os.path.dirname(__file__), "../../data/excel")
#     files    = glob.glob(os.path.join(data_dir, "*.xlsx"))
#     if not files:
#         raise FileNotFoundError(f"Không tìm thấy file xlsx trong {data_dir}")
#     for file_path in sorted(files, key=_sort_key):
#         print(f"📂 Đang chạy: {file_path}")
#         try:
#             save_disbursement(file_path)
#         except Exception as e:
#             print(f"⚠️ Lỗi {file_path}: {e} — upsert null")
#             try:
#                 time = parse_time_from_filename(file_path)
#                 df = pd.DataFrame([{"time": time}])
#                 upsert_disbursement(df)
#             except Exception as e2:
#                 print(f"⚠️ Bỏ qua {file_path}: {e2}")
        

##===============================================================

# if __name__ == "__main__":
#     disbursement()