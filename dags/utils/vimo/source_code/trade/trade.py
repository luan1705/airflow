import pandas as pd
import re
import os
import glob
from datetime import date
from sqlalchemy import create_engine, text
import openpyxl

DB_URL = "postgresql+psycopg2://root:Dnl_123456@tanhungsoft.com:5432/dnl"
engine = create_engine(DB_URL)

SCHEMA = "macro"
TABLE  = "trade"

COLS = [
    "totalTradeValue",
    "exportValue",
    "importValue",
    "tradeBalance",
    "tradeBalance_ytd",
    "exportDomestic",
    "exportForeign",
    "importDomestic",
    "importForeign",
    "tradeBalanceDomestic",
    "tradeBalanceForeign",
    "tradeBalanceDomestic_ytd",
    "tradeBalanceForeign_ytd",
    "exportForeignRatio",
]


def parse_time_from_filename(file_path: str) -> date:
    name = file_path.split("/")[-1].split("\\")[-1]
    match = re.search(r'(\d{4})_(\d{2})', name)
    if not match:
        raise ValueError(f"Không parse được tháng/năm từ tên file: {name}")
    return date(int(match.group(1)), int(match.group(2)), 1)


def _sheet_content(file_path, sheet, max_row=3):
    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    ws = wb[sheet]
    content = ' '.join([str(v) for row in ws.iter_rows(max_row=max_row, values_only=True) for v in row if v])
    wb.close()
    return content.lower()


def get_xk_sheet(file_path: str) -> str:
    xl = pd.ExcelFile(file_path)
    for sheet in xl.sheet_names:
        s = sheet.lower().replace(' ', '').replace('.', '')
        skip = ['nk', 'ldcn', 'lđcn', 'quy', 'quý', 'dichvu', 'dịchvụ', 'giaxk']
        if ('xk' in s or 'xuatkhau' in s or 'xuấtkhẩu' in s) and not any(k in s for k in skip):
            return sheet
    for sheet in xl.sheet_names:
        content = _sheet_content(file_path, sheet)
        if 'hàng hóa xuất khẩu' in content and 'tổng trị giá' in content:
            return sheet
    raise ValueError(f"Không tìm thấy sheet XK trong {file_path}")


def get_nk_sheet(file_path: str) -> str:
    xl = pd.ExcelFile(file_path)
    for sheet in xl.sheet_names:
        s = sheet.lower().replace(' ', '').replace('.', '')
        skip = ['xk', 'ldcn', 'lđcn', 'quy', 'quý', 'dichvu', 'dịchvụ', 'gianvl', 'giank']
        if ('nk' in s or 'nhapkhau' in s or 'nhậpkhẩu' in s) and not any(k in s for k in skip):
            return sheet
    for sheet in xl.sheet_names:
        content = _sheet_content(file_path, sheet)
        if 'hàng hóa nhập khẩu' in content and 'tổng trị giá' in content:
            return sheet
    raise ValueError(f"Không tìm thấy sheet NK trong {file_path}")


def to_float(val):
    try:
        return float(val)
    except (TypeError, ValueError):
        return None


def get_value_col(ws, year: int):
    col_texts = {}
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if any(v is not None for v in row):
            for ci, val in enumerate(row):
                if val is not None:
                    col_texts[ci] = col_texts.get(ci, '') + ' ' + str(val).strip().lower()
        if i > 10:
            break

    month_col = None
    for ci, text in sorted(col_texts.items()):
        if ci <= 0:
            continue
        if str(year) in text and 'tháng' in text and 'so với' not in text and 'cùng kỳ' not in text:
            if not re.search(r'[0-9]+ tháng', text):
                month_col = ci
                break

    if month_col is None:
        return None

    for ci in range(month_col, month_col + 3):
        if ci in col_texts and 'trị giá' in col_texts[ci]:
            return ci

    return month_col + 1


def parse_sheet(ws, val_col) -> dict:
    result = {'total': None, 'domestic': None, 'foreign': None}

    for row in ws.iter_rows(values_only=True):
        name = None
        for ci in [0, 1]:
            if len(row) > ci and row[ci] and str(row[ci]).strip():
                name = str(row[ci]).strip()
                break
        if not name:
            continue

        val = to_float(row[val_col] if val_col is not None and len(row) > val_col else None)

        if 'TỔNG TRỊ GIÁ' in name and result['total'] is None:
            result['total'] = val
        elif 'trong nước' in name.lower() and result['domestic'] is None:
            result['domestic'] = val
        elif ('vốn đầu tư' in name.lower() or 'nước ngoài' in name.lower() or 'đầu tư nn' in name.lower()) and result['foreign'] is None:
            result['foreign'] = val

        if all(v is not None for v in result.values()):
            break

    return result


def parse_trade(file_path: str) -> pd.DataFrame:
    time = parse_time_from_filename(file_path)
    year = time.year
    result = {"time": time}
    for col in COLS:
        result[col] = None

    # XK
    sheet_xk = get_xk_sheet(file_path)
    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    ws = wb[sheet_xk]
    val_col = get_value_col(ws, year)
    xk = parse_sheet(ws, val_col)
    wb.close()

    result['exportValue']    = round(xk['total'] * 1_000_000) if xk['total'] else None
    result['exportDomestic'] = round(xk['domestic'] * 1_000_000) if xk['domestic'] else None
    result['exportForeign']  = round(xk['foreign'] * 1_000_000) if xk['foreign'] else None

    # NK
    sheet_nk = get_nk_sheet(file_path)
    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    ws = wb[sheet_nk]
    val_col = get_value_col(ws, year)
    nk = parse_sheet(ws, val_col)
    wb.close()

    result['importValue']    = round(nk['total'] * 1_000_000) if nk['total'] else None
    result['importDomestic'] = round(nk['domestic'] * 1_000_000) if nk['domestic'] else None
    result['importForeign']  = round(nk['foreign'] * 1_000_000) if nk['foreign'] else None

    # Cán cân
    result['tradeBalance'] = (
        round(((xk['total'] or 0) - (nk['total'] or 0)) * 1_000_000)
        if xk['total'] is not None or nk['total'] is not None else None
    )
    result['tradeBalanceDomestic'] = (
        round(((xk['domestic'] or 0) - (nk['domestic'] or 0)) * 1_000_000)
        if xk['domestic'] is not None or nk['domestic'] is not None else None
    )
    result['tradeBalanceForeign'] = (
        round(((xk['foreign'] or 0) - (nk['foreign'] or 0)) * 1_000_000)
        if xk['foreign'] is not None or nk['foreign'] is not None else None
    )
    result['totalTradeValue'] = (
        round(((xk['total'] or 0) + (nk['total'] or 0)) * 1_000_000)
        if xk['total'] is not None or nk['total'] is not None else None
    )

    dom = result['exportDomestic']
    for_ = result['exportForeign']
    result['exportForeignRatio'] = (
        for_ / (dom + for_)
        if dom is not None and for_ is not None and (dom + for_) != 0 else None
    )

    return pd.DataFrame([result])


def upsert_trade(df: pd.DataFrame):
    upsert_cols = [c for c in COLS if '_ytd' not in c]
    with engine.begin() as conn:
        conn.execute(text(f'CREATE SCHEMA IF NOT EXISTS "{SCHEMA}"'))
        col_defs  = "\n".join([f'    "{c}" DOUBLE PRECISION,' for c in COLS[:-1]])
        col_defs += f'\n    "{COLS[-1]}" DOUBLE PRECISION'
        conn.execute(text(f"""
            CREATE TABLE IF NOT EXISTS {SCHEMA}.{TABLE} (
                time DATE PRIMARY KEY,
{col_defs}
            )
        """))
        for col in COLS:
            conn.execute(text(f"""
                ALTER TABLE {SCHEMA}.{TABLE}
                ADD COLUMN IF NOT EXISTS "{col}" DOUBLE PRECISION
            """))

        set_clause  = ",\n".join([f'    "{c}" = EXCLUDED."{c}"' for c in upsert_cols])
        insert_cols = ", ".join([f'"{c}"' for c in upsert_cols])
        insert_vals = ", ".join([f":{c}" for c in upsert_cols])
        records = df[['time'] + upsert_cols].replace({float('nan'): None}).to_dict(orient="records")
        conn.execute(text(f"""
            INSERT INTO {SCHEMA}.{TABLE} (time, {insert_cols})
            VALUES (:time, {insert_vals})
            ON CONFLICT (time) DO UPDATE SET
{set_clause}
        """), records)

    print(f"✅ Upsert {len(df)} rows vào {SCHEMA}.{TABLE}")


def save_trade(file_path: str):
    df = parse_trade(file_path)
    print(df.T.to_string())
    upsert_trade(df)


def _sort_key(f):
    match = re.search(r'(\d{4})_(\d{2})', os.path.basename(f))
    return (int(match.group(1)), int(match.group(2))) if match else (0, 0)


def get_latest_file(data_dir: str) -> str:
    files = glob.glob(os.path.join(data_dir, "*.xlsx"))
    if not files:
        raise FileNotFoundError(f"Không tìm thấy file xlsx trong {data_dir}")
    return sorted(files, key=_sort_key)[-1]


#=======================Chạy file chỉ định trực tiếp trong terminal=====================
# def trade(**context):
#     save_trade("../../data/2026_01.xlsx")

#=======================Chạy file chỉ định airflow=====================
# def trade(**context):
#     save_trade("/opt/airflow/dags/utils/vimo/data/2026_01.xlsx")

# =====================Chạy file mới nhất=====================
# def trade(**context):
#     data_dir  = os.path.join(os.path.dirname(__file__), "../../data")
#     file_path = get_latest_file(data_dir)
#     print(f"📂 File mới nhất: {file_path}")
#     save_trade(file_path)

# =====================Chạy tất cả file=====================
def trade(**context):
    data_dir = os.path.join(os.path.dirname(__file__), "../../data")
    files    = glob.glob(os.path.join(data_dir, "*.xlsx"))
    if not files:
        raise FileNotFoundError(f"Không tìm thấy file xlsx trong {data_dir}")
    for file_path in sorted(files, key=_sort_key):
        print(f"📂 Đang chạy: {file_path}")
        try:
            save_trade(file_path)
        except Exception as e:
            print(f"⚠️ Lỗi {file_path}: {e} — upsert null")
            try:
                time = parse_time_from_filename(file_path)
                df = pd.DataFrame([{"time": time}])
                upsert_trade(df)
            except Exception as e2:
                print(f"⚠️ Bỏ qua {file_path}: {e2}")

#===============================================================

if __name__ == "__main__":
    trade()