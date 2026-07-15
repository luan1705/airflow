import pandas as pd
import re
import os
import glob
from datetime import date
from sqlalchemy import create_engine, text
import openpyxl

DB_URL = "postgresql+psycopg2://vnsfintech:Vns_123456@tanhungsoft.com:5433/vnsfintech"
engine = create_engine(DB_URL)

SCHEMA = "macro"
TABLE  = "iip"

COMPONENTS = {
    "Toàn ngành công nghiệp":                                                          "total",
    "Khai khoáng":                                                                     "mining",
    "Công nghiệp chế biến, chế tạo":                                                  "manufacturing",
    "Sản xuất và phân phối điện":                                                      "electricity",
    "Sản xuất và phân phối điện, khí đốt, nước nóng, hơi nước và điều hoà không khí": "electricity",
}


def parse_time_from_filename(file_path: str) -> date:
    name = file_path.split("/")[-1].split("\\")[-1]
    match = re.search(r'(\d{4})_(\d{2})', name)
    if not match:
        raise ValueError(f"Không parse được tháng/năm từ tên file: {name}")
    return date(int(match.group(1)), int(match.group(2)), 1)


def get_sheet_name(file_path: str) -> str:
    xl = pd.ExcelFile(file_path)
    for sheet in xl.sheet_names:
        s = sheet.lower().replace(' ', '').replace('.', '')
        if 'iip' in s and 'thang' in s:
            return sheet
    for sheet in xl.sheet_names:
        s = sheet.lower().replace(' ', '').replace('.', '')
        if 'iip' in s:
            return sheet
    for sheet in xl.sheet_names:
        if re.match(r'^\s*2[\.\s]', sheet):
            return sheet
    raise ValueError(f"Không tìm thấy sheet IIP trong {file_path}")


def to_float(val):
    try:
        return float(val)
    except (TypeError, ValueError):
        return None


def get_yoy_col(ws, year: int):
    col_texts = {}
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if any(v is not None for v in row):
            for ci, val in enumerate(row):
                if val is not None:
                    col_texts[ci] = col_texts.get(ci, '') + ' ' + str(val).strip().lower()
        if i > 10:
            break

    yoy_cols = []
    for ci, text in sorted(col_texts.items()):
        if ci <= 0:
            continue
        if str(year) in text and 'cùng kỳ' in text and 'so với' in text:
            if not re.search(r'[0-9]+ tháng', text) and 'quý' not in text:
                yoy_cols.append(ci)

    # Lấy cột cuối = YoY tháng hiện tại
    return yoy_cols[-1] if yoy_cols else None


def parse_iip(file_path: str) -> pd.DataFrame:
    time  = parse_time_from_filename(file_path)
    year  = time.year
    sheet = get_sheet_name(file_path)

    wb      = openpyxl.load_workbook(file_path, read_only=True)
    ws      = wb[sheet]
    yoy_col = get_yoy_col(ws, year)

    unique_cols = list(dict.fromkeys(COMPONENTS.values()))
    result = {"time": time}
    for col in unique_cols:
        result[col] = None

    if yoy_col is None:
        wb.close()
        return pd.DataFrame([result])

    for row in ws.iter_rows(values_only=True):
        if not row[0] or not str(row[0]).strip():
            continue
        name = str(row[0]).strip().replace('\n', ' ')
        # Normalize tên (bỏ xuống dòng)
        name_norm = ' '.join(name.split())
        match_name = None
        for comp_name in COMPONENTS:
            if name_norm.startswith(' '.join(comp_name.split())):
                match_name = comp_name
                break

        if not match_name:
            continue

        col = COMPONENTS[match_name]
        if result[col] is not None:
            continue

        val = to_float(row[yoy_col] if len(row) > yoy_col else None)
        result[col] = round((val - 100) / 100, 4) if val is not None else None

    wb.close()
    result['totalAvg3m'] = None
    return pd.DataFrame([result])


def upsert_iip(df: pd.DataFrame):
    cols = list(dict.fromkeys(COMPONENTS.values())) + ['totalAvg3m']

    with engine.begin() as conn:
        conn.execute(text(f'CREATE SCHEMA IF NOT EXISTS "{SCHEMA}"'))
        col_defs  = "\n".join([f'    "{c}" DOUBLE PRECISION,' for c in cols[:-1]])
        col_defs += f'\n    "{cols[-1]}" DOUBLE PRECISION'
        conn.execute(text(f"""
            CREATE TABLE IF NOT EXISTS {SCHEMA}.{TABLE} (
                time DATE PRIMARY KEY,
{col_defs}
            )
        """))
        for col in cols:
            conn.execute(text(f"""
                ALTER TABLE {SCHEMA}.{TABLE}
                ADD COLUMN IF NOT EXISTS "{col}" DOUBLE PRECISION
            """))

        set_clause  = ",\n".join([f'    "{c}" = EXCLUDED."{c}"' for c in cols])
        insert_cols = ", ".join([f'"{c}"' for c in cols])
        insert_vals = ", ".join([f":{c}" for c in cols])
        records = df.replace({float('nan'): None}).to_dict(orient="records")
        conn.execute(text(f"""
            INSERT INTO {SCHEMA}.{TABLE} (time, {insert_cols})
            VALUES (:time, {insert_vals})
            ON CONFLICT (time) DO UPDATE SET
{set_clause}
        """), records)

    print(f"✅ Upsert {len(df)} rows vào {SCHEMA}.{TABLE}")


def save_iip(file_path: str):
    df = parse_iip(file_path)
    print(df.T.to_string())
    upsert_iip(df)


def _sort_key(f):
    match = re.search(r'(\d{4})_(\d{2})', os.path.basename(f))
    return (int(match.group(1)), int(match.group(2))) if match else (0, 0)


def get_latest_file(data_dir: str) -> str:
    files = glob.glob(os.path.join(data_dir, "*.xlsx"))
    if not files:
        raise FileNotFoundError(f"Không tìm thấy file xlsx trong {data_dir}")
    return sorted(files, key=_sort_key)[-1]


#=======================Chạy file chỉ định trực tiếp trong terminal=====================
# def iip(**context):
#     save_iip("../../data/2026_01.xlsx")

#=======================Chạy file chỉ định airflow=====================
# def iip(**context):
#     save_iip("/opt/airflow/dags/utils/vimo/data/2026_01.xlsx")

# =====================Chạy file mới nhất=====================
# def iip(**context):
#     data_dir  = os.path.join(os.path.dirname(__file__), "../../data")
#     file_path = get_latest_file(data_dir)
#     print(f"📂 File mới nhất: {file_path}")
#     save_iip(file_path)

# =====================Chạy tất cả file=====================
def iip(**context):
    data_dir = os.path.join(os.path.dirname(__file__), "../../data")
    files    = glob.glob(os.path.join(data_dir, "*.xlsx"))
    if not files:
        raise FileNotFoundError(f"Không tìm thấy file xlsx trong {data_dir}")
    for file_path in sorted(files, key=_sort_key):
        print(f"📂 Đang chạy: {file_path}")
        save_iip(file_path)

#===============================================================

if __name__ == "__main__":
    iip()